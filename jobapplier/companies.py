"""Load the target-company list out of the visa-sponsor spreadsheet."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable

import openpyxl

from .models import Company

# Header labels we understand, normalised to lowercase. The sheet's own header
# row is matched against these so column order changes don't break the parse.
_COLUMN_ALIASES = {
    "company": "name",
    "company name": "name",
    "employer": "name",
    "sector": "sector",
    "industry": "sector",
    "linkedin": "linkedin",
    "official job portal": "portal",
    "job portal": "portal",
    "careers page": "portal",
    "career site": "portal",
    "url": "portal",
    "visa sponsorship evidence": "sponsorship_evidence",
    "sponsorship evidence": "sponsorship_evidence",
    "what to verify": "notes",
    "notes": "notes",
}


def _find_header_row(ws, max_scan: int = 25) -> tuple[int, dict[int, str]]:
    """Locate the header row and map column index -> Company attribute name.

    The sheet has title/subtitle banner rows above the real header, so we scan
    down until we find a row that names at least a company and a portal column.
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan), start=1):
        mapping: dict[int, str] = {}
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            key = cell.value.strip().lower()
            if key in _COLUMN_ALIASES:
                mapping[cell.column] = _COLUMN_ALIASES[key]
        if "name" in mapping.values() and "portal" in mapping.values():
            return row_idx, mapping
    raise ValueError(
        f"No header row with both a company and a job-portal column found in sheet {ws.title!r}"
    )


def load_companies(xlsx_path: str | Path, sheet: str | None = None) -> list[Company]:
    """Parse the spreadsheet into Company records.

    Picks the sheet with the most data rows when none is named, so an
    Overview/Read-Me tab never gets mistaken for the data tab.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if sheet:
        candidates: Iterable = [wb[sheet]]
    else:
        candidates = wb.worksheets

    best: list[Company] = []
    for ws in candidates:
        try:
            header_row, colmap = _find_header_row(ws)
        except ValueError:
            continue

        rows: list[Company] = []
        for row in ws.iter_rows(min_row=header_row + 1):
            values: dict[str, str] = {}
            for cell in row:
                attr = colmap.get(cell.column)
                if not attr:
                    continue
                # Prefer a real hyperlink target over display text - the sheet
                # sometimes shows a shortened label but links the full URL.
                if cell.hyperlink and cell.hyperlink.target and attr in ("portal", "linkedin"):
                    values[attr] = cell.hyperlink.target.strip()
                elif cell.value is not None:
                    values[attr] = str(cell.value).strip()

            name = values.get("name", "")
            if not name or not values.get("portal"):
                continue
            rows.append(Company(**values))

        if len(rows) > len(best):
            best = rows

    if not best:
        raise ValueError(f"No company rows parsed from {xlsx_path}")
    return best


def save_companies(companies: list[Company], out_path: str | Path) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps([c.to_dict() for c in companies], indent=2))
    return out


def load_companies_json(path: str | Path) -> list[Company]:
    data = json.loads(Path(path).read_text())
    return [Company(**c) for c in data]
